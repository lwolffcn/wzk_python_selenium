# coding=utf-8
import unittest, random, os, sys
from time import sleep
from pages.login_page import Loginpage
from models import myunit, function
from tkinter import Tk
from tkinter.messagebox import showwarning
import win32com.client as win32
from time import sleep
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl, datetime
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference, Series


def test_file():
    filename = r'D:\wzk_python_selenium\PC4.0\files\test_input.txt'
    # f = open(filename ,'r')
    # print(f.read())
    # f.close()
    with open(filename) as f:
        for line in f:
            print(line)

warn = lambda app: showwarning(app,'Exit?')
RANGE = range(1,8)

def excel():
    app = 'Excel'
    #xl = win32.gencache.EnsureDispatch('%s.Application' % app)
    xl = win32.Dispatch('%s.Application' % app)
    ss = xl.Workbooks.Add()
    sh = ss.ActiveSheet
    xl.Visible = True
    sleep(1)
    sh.Cells(1,1).Value = 'dasasdasd'
    warn(app)
    ss.Close(True)
    xl.Application.Quit()

def oprnexcel():
    filename = r'D:\wzk_python_selenium\PC4.0\files\yesOrNo.xlsx'
    # wb = Workbook()
    # 激活 worksheet
    # ws = wb.active # 默认设置为0。除非修改其值，否则使用此方法将始终获取第一个工作表。
    inwb = openpyxl.load_workbook(filename)
    sheets = inwb.sheetnames
    print(inwb[sheets[0]])
    print(inwb[sheets[-1]])
    ws = inwb['真假题']
    rows = ws.max_row
    cols = ws.max_column
    print(rows)
    print(cols)
    for i in range(1, cols + 1):
        print(ws.cell(1, i).value)

    outwb = Workbook()
    first_sheet = outwb['Sheet']
    first_sheet.title = 'first_sheet'
    outws = outwb.create_sheet('111', index = 0)
    outws.sheet_properties.tabColor = "1072BA"
    outws['A1'] = 42
    outws['A2'] = datetime.datetime.now()
    outws2 = outwb.create_sheet('222', index=1)
    outws3 = outwb.create_sheet('333', index=2)
    outws4 = outwb.create_sheet('444', index=3)
    outws5 = outwb.create_sheet('555', index=-1)
    outws.append([1, 2, 3])
    # for row in range(1, 70):
    #     for col in range(1, 4):
    #         outws.cell(row, col).value = row * 2  # 写文件
    #     print(row)
    saveExecl =  r'D:\wzk_python_selenium\PC4.0\files\yesOrNo2.xlsx'
    # outwb.save(saveExecl)

    write_only_file = r'D:\wzk_python_selenium\PC4.0\files\write_only_file.xlsx'
    wbwb = Workbook(write_only=True)
    wsws = wbwb.create_sheet()
    cell = WriteOnlyCell(wsws, value="hello world")
    cell.font = Font(name='Courier', size=36)
    cell.comment = Comment(text="A comment", author="Author's Name")
    wsws.append([cell, 3.14, None])
    wbwb.save(write_only_file)
    '''
    与普通工作簿不同，新创建的只写工作簿不包含任何工作表；必须使用 create_sheet() 方法。
    在只写工作簿中，只能使用 append() . 不能在任意位置用 cell() 或 iter_rows() .
    它能够导出无限量的数据（甚至超过了Excel的实际处理能力），同时将内存使用量保持在10MB以下。
    只写工作簿只能保存一次。之后，每次试图将工作簿或append（）保存到现有工作表时，都会引发 openpyxl.utils.exceptions.WorkbookAlreadySaved 例外。
    在添加单元格之前，必须创建实际单元格数据之前出现在文件中的所有内容，因为在此之前必须将其写入文件。例如， freeze_panes 应在添加单元格之前设置。
    '''
    SampleChart = r'D:\wzk_python_selenium\PC4.0\files\SampleChart.xlsx'
    wbSampleChart = Workbook()
    wsSampleChart = wbSampleChart.active
    for i in range(10):
        wsSampleChart.append([i])
    values = Reference(wsSampleChart, min_col=1, min_row=1, max_col=1, max_row=10)
    chart = BarChart()
    chart.add_data(values)
    wsSampleChart.add_chart(chart, "E15")
    wbSampleChart.save(SampleChart)



if __name__=='__main__':
    # Tk().withdraw()
    # excel()
    oprnexcel()

